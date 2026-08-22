#!/usr/bin/env python3
"""Generate the family chores spreadsheet from whatever is live right now.

Upload to Google Drive -> it converts to a Sheet -> File > Share > Publish to web
-> pick the Chores tab -> CSV -> paste that link into the dashboard's
"Spreadsheet link..." button. Column headers are matched by keyword, so they can be
reworded freely.
"""
import json
import os
import subprocess
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Family Chores.xlsx")
API = "http://192.168.1.226:5000/chores"

HEADERS = ["Chore Title", "Points", "Chore Steps/Details",
           "Type (Required / Optional)", "Frequency"]
NAVY, STEEL, INK = "1B2A4A", "4A78D6", "222222"


def live_chores():
    """Pull the current catalog through the HA box (the API is LAN-only)."""
    try:
        raw = subprocess.check_output([
            "ssh", "-i", os.path.join(HERE, "keys", "id_ha"),
            "-o", "StrictHostKeyChecking=no", "root@192.168.1.226",
            "curl -s -m 20 %s" % API], text=True, encoding="utf-8", timeout=60)
        return json.loads(raw).get("chores") or []
    except Exception as exc:
        print("could not reach the chores API (%s); writing headers only" % exc)
        return []


def style_header(ws, headers):
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


chores = sorted(live_chores(), key=lambda c: (c.get("kind") != "required",
                                              {"weekly": 0, "bi-weekly": 1, "monthly": 2}
                                              .get(c.get("frequency"), 9),
                                              c.get("name", "")))

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- Chores tab
ws = wb.active
ws.title = "Chores"
style_header(ws, HEADERS)
for c in chores:
    ws.append([c.get("name", ""), int(c.get("points", 0)), c.get("description", ""),
               c.get("kind", "required"), c.get("frequency", "weekly")])
for _ in range(15):
    ws.append(["", None, "", "", ""])

dv_kind = DataValidation(type="list", formula1='"required,optional"', allow_blank=True)
dv_kind.promptTitle = "Chore type"
dv_kind.prompt = ("required = has to be done before anyone can claim an optional one.\n"
                  "optional = extra points, unlocks once every required chore is claimed.")
ws.add_data_validation(dv_kind)
dv_kind.add("D2:D200")

dv_freq = DataValidation(type="list", formula1='"daily,weekly,bi-weekly,monthly"', allow_blank=True)
dv_freq.promptTitle = "How often"
dv_freq.prompt = ("How long before this chore comes back after it's been done.\n"
                  "daily = every day, weekly = 7 days, bi-weekly = 14 days, monthly = 30 days.")
ws.add_data_validation(dv_freq)
dv_freq.add("E2:E200")

for col, w in {"A": 36, "B": 9, "C": 74, "D": 20, "E": 14}.items():
    ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=5):
    row[1].alignment = Alignment(horizontal="center")
    row[2].alignment = Alignment(wrap_text=True, vertical="top")
    row[3].alignment = Alignment(horizontal="center")
    row[4].alignment = Alignment(horizontal="center")

# ---------------------------------------------------------------- Rotation tab
rot = wb.create_sheet("Rotation")
style_header(rot, ["Chore Title", "Frequency", "Last Done", "Next Time It Posts"])
for c in chores:
    rot.append([c.get("name", ""), c.get("frequency", "weekly"),
                c.get("last_done") or "— never —", c.get("next_due") or "on the board now"])
for col, w in {"A": 36, "B": 14, "C": 16, "D": 20}.items():
    rot.column_dimensions[col].width = w
for row in rot.iter_rows(min_row=2, max_row=rot.max_row, max_col=4):
    for cell in row[1:]:
        cell.alignment = Alignment(horizontal="center")

note = rot.cell(row=rot.max_row + 2, column=1,
                value="Read-only snapshot. The dashboard tracks these dates live as the "
                      "boys claim chores — it can't write back into this sheet.")
note.font = Font(italic=True, color="777777", size=10)

# ---------------------------------------------------------------- guide tab
gd = wb.create_sheet("How this works")
gd.column_dimensions["A"].width = 112
notes = [
    ("Family chores — how this sheet is used", True),
    ("", False),
    ("Anything you type on the 'Chores' tab shows up on the wall tablet within 5 minutes.", False),
    ("", False),
    ("Chore Title                  the name the boys will see", False),
    ("Points                       the points it's worth (a whole number)", False),
    ("Chore Steps/Details          how to do it — shows under the chore name on the tablet", False),
    ("Type (Required / Optional)   pick from the dropdown", False),
    ("Frequency                    weekly / bi-weekly / monthly — pick from the dropdown", False),
    ("", False),
    ("You can rename these headers. Columns are matched by keyword, so 'Task'/'Job'/", False),
    ("'Title' work for the name, 'Pts'/'Value' for points, 'Instructions'/'How to' for", False),
    ("the steps, 'Kind'/'Category' for the type, and 'How Often'/'Repeat' for frequency.", False),
    ("", False),
    ("Required vs optional", True),
    ("Every REQUIRED chore on the board has to be claimed before any OPTIONAL one can", False),
    ("be. Optional chores show with a lock until then, so the boys can't skip ahead to", False),
    ("the fun high-point jobs.", False),
    ("", False),
    ("How the frequency rotation works", True),
    ("The board is rebuilt every Monday. A chore only goes back up once its interval", False),
    ("has passed since it was last finished:", False),
    ("    weekly      back on the board 7 days after it was done", False),
    ("    bi-weekly   back after 14 days", False),
    ("    monthly     back after 30 days", False),
    ("So a monthly job finished on the 3rd won't reappear until the 2nd of next month.", False),
    ("A chore nobody finished just stays up — it doesn't get skipped.", False),
    ("The 'Rotation' tab shows where every chore currently sits.", False),
    ("", False),
    ("Three places to add chores — they all stay in sync", True),
    ("This sheet, the Cozi lists 'Chores Required' / 'Chores Optional', and the tablet", False),
    ("itself. A chore added in any one appears in the other two. Matching is by name,", False),
    ("so the same chore never shows up twice.", False),
    ("", False),
    ("This sheet is the source of truth for FREQUENCY — Cozi list items can't carry", False),
    ("one, so change frequency here (or on the tablet), not in Cozi.", False),
    ("", False),
    ("Deleting a row here removes the chore from the tablet too — unless one of the", False),
    ("boys already claimed it this week, which is protected.", False),
]
for i, (text, bold) in enumerate(notes, start=1):
    c = gd.cell(row=i, column=1, value=text)
    c.font = Font(bold=bold, size=12 if bold else 11, color=STEEL if bold else INK)

wb.save(OUT)
print("wrote %s (%d bytes) — %d chores across %d tabs"
      % (OUT, os.path.getsize(OUT), len(chores), len(wb.sheetnames)))
